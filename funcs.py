from datetime import date, timedelta
from openpyxl import load_workbook
from tkinter import *

root = Tk()
root.title("Планирование финансов")
root.geometry("{0}x{1}+0+0".format(root.winfo_screenwidth(), root.winfo_screenheight()))

plan_wb = load_workbook("categories.xlsx")
finance_plan = plan_wb.active
fact_wb = load_workbook("document.xlsx")
finance_fact = fact_wb.active
frame1 = Frame()
# warn = Label(text="Введённое значение должно быть положительным числом")



def check_text(text):
    try:
        if int(text) > 0:
            return int(text)
        else:
            return False
    except:
        return False

def get_zero(num):
    if num < 10:
        result = '0' + str(num)
    else:
        result = str(num)
    return result

def get_week():
    today = date.today()
    first_day = today - timedelta(today.weekday())
    last_day = first_day + timedelta(7)
    return get_zero(first_day.day)+ "." + get_zero(first_day.month) + " - " + get_zero(last_day.day) + "." + get_zero(last_day.month)

def enter_categories(temp, categories, values, frame):
    wb = load_workbook("categories.xlsx")
    ws = wb.active
    i = 1
    for el in range (len(temp)):
        # print(categories[el], temp[el].get(), values[el].get())
        if temp[el].get() == 1:
            ws.cell(1, i, value=categories[el])
            val = values[el].get()
            try:
                val = int(val)
                ws.cell(2, i, value=val)
            except:
                Label(frame, text="Значения не должны содержать буквы").grid()
                return 1
            i += 1
    frame.destroy()
    print("works")
    wb.save("categories.xlsx")

def get_all_categories():
    return open("list of categories.txt", "r", encoding="utf-8").read().split(", ")

def get_chosen_categories():
    categories = []
    values = []
    wb = load_workbook("categories.xlsx")
    ws = wb.active
    i = 1
    while True:
        val = ws.cell(row=1, column=i).value
        i += 1
        if val == None:
            break
        else:
            categories.append(val)
            values.append(ws.cell(row=2, column=i-1).value)
    return [categories, values]

def get_statistics():
    canvas = Canvas(bg="white", width=350, height=300)
    canvas.grid()
    canvas.create_oval(25, 10, 45, 30, fill="blue")
    canvas.create_text(50, 40, text=("Осталось"))
    canvas.create_oval(205, 10, 225, 30, fill="red")
    canvas.create_text(200, 40, text=("Потрачено"))
    for i in range(len(get_chosen_categories()[0])):
        wb = load_workbook("document.xlsx")
        ws = wb.active
        canvas.create_text(50,70+30*i, text=(ws.cell(row=i+3, column=1).value))
        wasted = ws.cell(row=i+3, column=3).value
        left = ws.cell(row=i+3, column=2).value
        left_percent = left/(wasted + left)
        canvas.create_rectangle(100, 60+30*i, 100+200*left_percent, 80+30*i, fill="blue")
        if left_percent > 0.1:
            canvas.create_text(100+100*left_percent, 70+30*i, text=(str(left)))
        # rect_center_x = 
        canvas.create_rectangle(100+200*left_percent, 60+30*i, 325, 80+30*i, fill="red")
        if left_percent < 0.9:
           canvas.create_text(225+100*left_percent, 70+30*i, text=(str(wasted)))

def change_expences_plan(root):
    def change_vals(entries,frame):
        warn = Label(frame, text="")
        warn.grid()
        for i in range(len(entries)):
            val = check_text(entries[i].get())
            if not val:
                warn = Label(frame, text="Введённое значение должно быть положительным числом")
                warn.grid()
                return 1
            else:
                finance_plan.cell(row=2, column=i+1, value=val)
            warn.grid_forget()
        plan_wb.save("categories.xlsx")
        frame.grid_forget()
    frame = Frame(root,background='pink', width=400, height=400)
    frame.grid(row=0, column=0, rowspan=2)
    frame.grid_propagate(0)
    [all_categories, all_values] = get_chosen_categories()
    # print(all_values)
    entries = []
    for i in range(len(all_categories)):
        label = Label(frame, text=all_categories[i])
        label.grid(row=i, column=0)
        entry = Entry(frame)
        entry.insert(index=0, string=all_values[i])
        entries.append(entry)
        entry.grid(row=i, column=1)
    btn = Button(frame, text="Подтвердить", command= lambda: change_vals(entries, frame))
    btn.grid(columnspan=2)
    btn = Button(frame, text="Отменить", command=frame.grid_forget)
    btn.grid()

def add_new_expence(root):
    frame = Frame(root, background='pink', width=400, height=400)
    frame.grid(row=0, column=0, rowspan=2)
    frame.grid_propagate(0)
    def enter_expences():
        frame.grid_forget()
        return 0
    entries = []
    for i in range((len(get_chosen_categories()[0]))):
        label = Label(frame, text=get_chosen_categories()[0][i])
        label.grid(row=i, column=0)
        entry = Entry(frame)
        entry.grid(row=i, column=1)
        entries.append(entry)
    btn = Button(frame, text="Подтвердить", command=enter_expences)
    btn.grid()
    btn = Button(frame, text="Отменить", command=frame.grid_forget)
    btn.grid()
    return 0
def change_categories():
    return 0